#Экспорт данных в файл txt
import openpyxl

def export_data():
    wb = openpyxl.reader.excel.load_workbook(filename='dir_phone.xlsx')
    sheet = wb.active
    with open('dir_phone.txt','w', encoding='utf-8') as file:
        for i in range(2, sheet.max_row + 1):
            file.write(str(sheet['A' + str(i)].value)+" ")
            file.write(sheet['B' + str(i)].value+" ")
            file.write(str(sheet['C' + str(i)].value))
            file.write(f";\n")

def export_data_id(i):
    wb = openpyxl.reader.excel.load_workbook(filename='dir_phone.xlsx')
    sheet = wb.active
    with open('dir_phone.txt', 'a', encoding='utf-8') as file:
        with open('dir_phone.txt', 'r', encoding='utf-8') as file2:
            data = file2.read()
        for i in range(i+1, i + 2):
            if str(sheet['A' + str(i)].value) in data:
                print('Такой номер уже есть в файле')
            else:
                file.writelines(str(sheet['A' + str(i)].value) + " ")
                file.write(sheet['B' + str(i)].value + " ")
                file.write(str(sheet['C' + str(i)].value))
                file.write(f";\n")

