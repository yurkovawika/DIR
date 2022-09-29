#печать и поиск данных
import input_data as id
import openpyxl


def print_data_all():  # отображение всего справочника
    wb = openpyxl.reader.excel.load_workbook(filename='dir_phone.xlsx')
    sheet = wb.active
    for i in range(2, sheet.max_row + 1):
        print(i-1,'-',sheet['A' + str(i)].value, sheet['B' + str(i)].value, sheet['C' + str(i)].value)


def print_data_name():  # поиск по имени
    wb = openpyxl.reader.excel.load_workbook(filename='dir_phone.xlsx')
    sheet = wb.active
    name = id.name()
    found = False
    for i in range(2, sheet.max_row + 1):
        if name in sheet['A' + str(i)].value:
            found = True
            print(i - 1, '-', sheet['A' + str(i)].value, sheet['B' + str(i)].value, sheet['C' + str(i)].value)
    if not found:
        print("Поиск не дал результатов")


def print_data_last_name():  # поиск по фамилии
    wb = openpyxl.reader.excel.load_workbook(filename='dir_phone.xlsx')
    sheet = wb.active
    last_name = id.last_name()
    found = False
    for i in range(2, sheet.max_row + 1):
        if last_name in sheet['B' + str(i)].value:
            found = True
            print(i - 1, '-', sheet['A' + str(i)].value, sheet['B' + str(i)].value, sheet['C' + str(i)].value)
    if not found:
        print("Поиск не дал результатов")

def print_data_t_phone():  # поиск по номеру телефона
    wb = openpyxl.reader.excel.load_workbook(filename='dir_phone.xlsx')
    sheet = wb.active
    found = False
    t_phone = id.t_phone()
    for i in range(2, sheet.max_row + 1):
        if t_phone in sheet['C' + str(i)].value:
            print(i - 1, '-', sheet['A' + str(i)].value, sheet['B' + str(i)].value, sheet['C' + str(i)].value)
            found = True
    if not found:
        print("Поиск не дал результатов")
